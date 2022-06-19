from bs4 import BeautifulSoup #Biblioteca para achar tags html, depois vamos colocar num dataframe do pandas
import openpyxl #Biblioteca para criar a planilha do excel e editá-la
import pandas as pd #Biblioteca para guardar data frame
import yfinance as yf #Biblioteca do Yahoo Finance
import requests #Biblioteca para extrair uma URL
import qrcode #Biblioteca para gerar qrcode
from openpyxl.drawing.image import Image
from openpyxl.styles import (
    Alignment,
    PatternFill,
    Font,
    colors
) #Biblioteca para estilizar a tabela do excel

class Funcoes:
    #FUNÇÃO WEBSCRAPING DA CARTEIRA
    def achar_carteira_dic(url_carteira):
        # WebScraping
        # Url do html
        url = url_carteira
        # Obtendo o conteúdo da página em formato de texto

        data = requests.get(url).text
        soup = BeautifulSoup(data, "html.parser")

        # Procurando divs com classes:
        div_acao = soup.find("div", class_="acao")

        # Procurando as tabelas por classe na página
        table_class_acoes = div_acao.find("table")  # tabela de ações

        # Definindo dataframe
        # aqui por enquanto só definimos as colunas
        df_acoes = pd.DataFrame(columns=["Ação", "Quantidade"])  # dataframe da tabela ações

        # Contando o número de linhas em cada tabela
        linhas_na_tabela_acoes = []  # Aqui vamos por um item para linha da tabela com dados, para conseguir o número de linhas
        for row in table_class_acoes.find_all("tr"):  # em html uma linha da tabela é representada pela tag <tr>
            columns_acoes = row.find_all("td")  # em html uma coluna da tabela é representada pela tag <td>
            if (columns_acoes != []):
                linhas_na_tabela_acoes.append("linha")
        num_linhas_acoes = len(linhas_na_tabela_acoes)  # número de linhas em tabela ações
        # Obtendo todas as linhas da tabela
        # Obtendo todas as linhas da tabela acoes

        for row in table_class_acoes.find_all("tr"):  # em html uma linha da tabela é representada pela tag <tr>
            # Obtendo todas as colunas em cada linha
            columns_acoes = row.find_all("td")  # em html uma coluna da tabela é representada pela tag <td>

            if (columns_acoes != []):
                acao = columns_acoes[0].text.strip(" ")
                quantidade = columns_acoes[1].text.strip(" ")
                df_acoes = pd.concat([df_acoes, pd.DataFrame.from_records([{"Ação": acao, "Quantidade": quantidade}])],
                                     ignore_index=True)
                # Aqui a primeira linha vai conter o nosso primeiro df_acoes definido que contém Ação, Quantidade, ou seja os nomes das nossas colunas
                # a segunda parte dos nossos dados serão achados pelo método pandas.DataFrame.from_records
                # nele teremos uma lista é claro, pois em cada linha temos mais de uma informação
                # nesta lista queremos a ação e a quantidade o que a gente tem!
                # pois definimos acao e quantidade
                # Podemos usar o método com dicionário! a chave do dic é a coluna e o valor é o que a coluna vai receber, no caso acao e quantidade IZI
                # o ignore_index é para enumerar nossas linhas, note que ele não coloca numero em ação e quantidade, pois definimos esse dado como coluna

        df_acoes.head(num_linhas_acoes)  # Para retornar as num_linhas_acoes desejadas

        # Refatorando o dataframe
        # Aqui só vou arrumar os dados para ficar da forma certa, pois tratamos eles como str, mas não necessariamente são str
        # Ações
        # Ação
        df_acoes["Ação"] = df_acoes["Ação"].str.upper()
        # Quantidade
        df_acoes["Quantidade"] = [x.replace(',', '.') for x in df_acoes["Quantidade"]]
        df_acoes = df_acoes.astype({"Quantidade": float})

        df_acoes.head(num_linhas_acoes)

        # Procurando divs com classes:
        div_moeda = soup.find("div", class_="moeda")

        # Procurando as tabelas por classe na página
        table_class_moeda = div_moeda.find("table")  # tabela de moeda

        # Definindo dataframe
        # aqui por enquanto só definimos as colunas
        df_moedas = pd.DataFrame(columns=["Moeda", "Quantidade por tipo"])  # dataframe da tabela moeda

        # Contando o número de linhas em cada tabela
        linhas_na_tabela_moeda = []  # Aqui vamos por um item para linha da tabela com dados, para conseguir o número de linhas
        for row in table_class_moeda.find_all("tr"):  # em html uma linha da tabela é representada pela tag <tr>
            columns_moeda = row.find_all("td")  # em html uma coluna da tabela é representada pela tag <td>
            if (columns_moeda != []):
                linhas_na_tabela_moeda.append("linha")
        num_linhas_moeda = len(linhas_na_tabela_moeda)  # número de linhas em tabela ações
        # Obtendo todas as linhas da tabela
        # Obtendo todas as linhas da tabela acoes

        for row in table_class_moeda.find_all("tr"):  # em html uma linha da tabela é representada pela tag <tr>
            # Obtendo todas as colunas em cada linha
            columns_moeda = row.find_all("td")  # em html uma coluna da tabela é representada pela tag <td>

            if (columns_moeda != []):
                moeda = columns_moeda[0].text.strip(" ")
                quantidade_por_tipo = columns_moeda[1].text.strip(" ")
                df_moedas = pd.concat(
                    [df_moedas,
                     pd.DataFrame.from_records([{"Moeda": moeda, "Quantidade por tipo": quantidade_por_tipo}])],
                    ignore_index=True)
                # Aqui a primeira linha vai conter o nosso primeiro df_acoes definido que contém Moeda, Quantidade por tipo, ou seja os nomes das nossas colunas
                # a segunda parte dos nossos dados serão achados pelo método pandas.DataFrame.from_records
                # nele teremos uma lista é claro, pois em cada linha temos mais de uma informação
                # nesta lista queremos a moeda e a quantidade_por_tipo o que a gente tem!
                # pois definimos moeda e quantidade_por_tipo
                # Podemos usar o método com dicionário! a chave do dic é a coluna e o valor é o que a coluna vai receber, no caso moeda e quantidade_por_tipo IZI
                # o ignore_index é para enumerar nossas linhas, note que ele não coloca numero em moeda e quantidade por tipo, pois definimos esse dado como coluna

        df_moedas.head(num_linhas_moeda)  # Para retornar as num_linhas_acoes desejadas

        # Refatorando o dataframe
        # Aqui só vou arrumar os dados para ficar da forma certa, pois tratamos eles como str, mas não necessariamente são str
        # Ações
        # Ação
        df_moedas["Moeda"] = df_moedas["Moeda"].str.upper()
        # Quantidade
        df_moedas["Quantidade por tipo"] = [x.replace(',', '.') for x in df_moedas["Quantidade por tipo"]]
        df_moedas = df_moedas.astype({"Quantidade por tipo": float})

        df_moedas.head(num_linhas_moeda)

        dic_acao = {}
        cont_dic_acao = 0
        for acao in df_acoes["Ação"]:
            quantidade_acao = df_acoes["Quantidade"][cont_dic_acao]
            dic_acao[f"{acao}"] = f"{quantidade_acao}"
            cont_dic_acao += 1

        dic_moeda = {}
        cont_dic_moeda = 0
        for moeda in df_moedas["Moeda"]:
            quantidade_moeda = df_moedas["Quantidade por tipo"][cont_dic_moeda]
            dic_moeda[f"{moeda}"] = f"{quantidade_moeda}"
            cont_dic_moeda += 1
        dic_moeda_acao = {"Ação": dic_acao, "Moeda": dic_moeda}

        return [dic_moeda_acao, df_moedas, df_acoes]

    # função para tirar essa *** de BRL=X que mandaram deixar na carteira (era só mudar na função)
    def tire_brl_x(moeda):
        if moeda == "BRL":
            moeda = "BRL"
        elif moeda == "BRLBRL=X":
            moeda == "BRL"
        elif moeda == "BRL=X":
            moeda == "BRL"
        elif "BRL=X" in moeda is False:
            moeda = moeda
        else:
            # deixando normal a moeda
            letra_moeda = len(moeda)
            # tirando BRL=X
            num_letra_moeda = letra_moeda - 5
            # moeda sem BRL=X
            moeda = moeda[0:num_letra_moeda]
        return moeda

    #Função para coletar cotação das ações em reais
    def cotacao_acao(url_carteira):
        df_acao = Funcoes.achar_carteira_dic(url_carteira)[2]
        dic_acao_cot = {}
        for acao in df_acao["Ação"]:
            # checando se a ação existe
            if yf.Ticker(f"{acao}") == None:
                dic_acao_cot[f"{acao}"] = "Ação não encontrada"
            else:
                ticker = yf.Ticker(f"{acao}")
                # diz em que moeda está a ação
                moeda_acao = ticker.info['currency']
                # diz o valor da ação em sua moeda
                valor_acao = ticker.info["regularMarketPrice"]
                if moeda_acao != "BRL":
                    moeda_acao_real = moeda_acao + "BRL=X"
                    moeda_acao_real = yf.Ticker(moeda_acao_real)
                    # Valor da moeda da ação em real
                    valor_moeda_acao = moeda_acao_real.info["regularMarketPrice"]
                    # Valor da ação em real
                    valor_acao = valor_acao * valor_moeda_acao
                    dic_acao_cot[f"{acao}"] = valor_acao
                else:
                    valor_acao = ticker.info["regularMarketPrice"]
                    dic_acao_cot[f"{acao}"] = valor_acao
        return dic_acao_cot
    #Função para calcular valores das ações em reais
    def valor_total_acao(url_carteira):
        df_acao = Funcoes.achar_carteira_dic(url_carteira)[0]["Ação"]
        dic = Funcoes.cotacao_acao(url_carteira)
        dic_val_tot_acao = {}
        for k, v in dic.items():
            dic_val_tot_acao[k] = v * float(df_acao[k])
        return dic_val_tot_acao
    #Função para coletar cotação das moedas em reais
    def cotacao_moeda(url_carteira):
        df_moeda = Funcoes.achar_carteira_dic(url_carteira)[1]
        dic_moeda_cot = {}
        for moeda in df_moeda["Moeda"]:
            if yf.Ticker(f"{moeda}") == None:
                dic_moeda_cot[f"{moeda}"] = "Moeda não encontrada"
            elif moeda == "BRL=X":
                dic_moeda_cot["BRL"] = 1
            elif moeda == "BRL":
                dic_moeda_cot["BRL"] = 1
            elif moeda == "BRLBRL=X":
                dic_moeda_cot["BRL"] = 1
            if "BRL=X" in moeda is False:
                dic_moeda_cot[f"{moeda}"] = yf.Ticker(moeda + "BRL=X").info["regularMarketPrice"]
            else:
                # deixando normal a moeda
                letra_moeda = len(moeda)
                # tirando BRL=X
                num_letra_moeda = letra_moeda - 5
                # moeda sem BRL=X
                moeda_normal = moeda[0:num_letra_moeda]
                dic_moeda_cot[f"{moeda_normal}"] = yf.Ticker(moeda).info["regularMarketPrice"]

        return dic_moeda_cot
    #Função para coletar valor da moeda em reais
    def valor_total_moeda(url_carteira):
        df_moeda = Funcoes.achar_carteira_dic(url_carteira)[0]["Moeda"]
        # mudando chave do dicionário
        chaves_novas = {}
        for k, v in df_moeda.items():
            chaves_novas[k] = Funcoes.tire_brl_x(k)
            # Deculpa por usar esse negócio horrendo por favor não tire pontos
        df_moeda_normal = dict([(chaves_novas.get(key), value) for key, value in df_moeda.items()])
        dic = Funcoes.cotacao_moeda(url_carteira)
        dic_val_tot_moeda = {}
        for k, v in dic.items():
            dic_val_tot_moeda[k] = v * float(df_moeda_normal[k])
        return dic_val_tot_moeda

    def soma_valores(nome_arquivo):
        # Abre a planilha
        planilha = openpyxl.load_workbook(nome_arquivo)
        # Abre a folha Carteira
        carteira = planilha["Carteira"]
        # Pega o tamanho da carteira
        tamanho_tabela = len(list(carteira.rows))
        # Define o valor total da carteira
        soma_total = float()
        for linha in range(3, tamanho_tabela+1):
            soma_total += float(carteira.cell(row=linha, column=4).value)
            soma_total += float(carteira.cell(row=linha, column=8).value)
        # Salva a planilha
        return soma_total


    def criar_qrcode(nome_arquivo, nome_qrcode):
        # Puxa a função de cima
        soma_total = Funcoes.soma_valores(nome_arquivo)
        # Insere a mensagem que vai ser lida no qrcode
        carteira = (f"O valor da sua carteira é de R$ {round(soma_total, 2)}")
        # Ajusta o tamanho do qrcode
        estilo_qrcode = qrcode.QRCode(box_size=4, border=2)
        # Adiciona amensagem no qrcode
        estilo_qrcode.add_data(carteira)
        # Salva a imagem do qrcode
        imagem = estilo_qrcode.make_image()
        imagem.save(nome_qrcode)


    def colocar_qrcode(nome_arquivo):
        # dá nome ao qrcode e salva ele na pasta Imagens
        nome_qrcode = 'qrcode.png'
        Funcoes.criar_qrcode(nome_arquivo, nome_qrcode)
        # Abre a planilha
        planilha = openpyxl.load_workbook(nome_arquivo)
        # Abre a folha Carteira
        carteira = planilha["Carteira"]
        # Pega o tamanho da carteira
        tamanho_tabela = len(list(carteira.rows))
        # Coloca o qrcode aonde é definido
        linha_qrcode = str(tamanho_tabela + 3)
        celula = "D" + linha_qrcode
        imagem = Image(nome_qrcode)
        carteira.add_image(imagem, celula)
        # Salva a planilha
        planilha.save(nome_arquivo)

    #Função para gerar planilha no excel
    def gerar_planilha_carteira(url_carteira):
        df_acao = Funcoes.achar_carteira_dic(url_carteira)[2]
        df_moeda = Funcoes.achar_carteira_dic(url_carteira)[1]
        cot_acao = Funcoes.cotacao_acao(url_carteira)
        cot_moeda = Funcoes.cotacao_moeda(url_carteira)
        total_acao = Funcoes.valor_total_acao(url_carteira)
        total_moeda = Funcoes.valor_total_moeda(url_carteira)
        # Criando planilha
        # Criando folha da carteira
        wb = openpyxl.Workbook()
        sheet_1 = wb.active
        sheet_1.title = "Carteira"

        # tabela de ações
        sheet_1.merge_cells("A1:D1")
        cell_1 = sheet_1.cell(row=1, column=1)
        cell_1.value = "Ação"
        # Estilizando a tabela de ações
        cell_1.alignment = Alignment(horizontal="center", vertical="center")
        sheet_1["A2"] = "Ação"
        sheet_1["B2"] = "Quantidade"
        sheet_1["C2"] = "Cotação da ação (R$)"
        sheet_1["D2"] = "Valor atual total ação (R$)"

        sheet_1["A2"].fill = PatternFill(start_color="38b2eb", end_color="38b2eb", fill_type="solid")
        sheet_1["B2"].fill = PatternFill(start_color="38b2eb", end_color="38b2eb", fill_type="solid")
        sheet_1["C2"].fill = PatternFill(start_color="38b2eb", end_color="38b2eb", fill_type="solid")
        sheet_1["D2"].fill = PatternFill(start_color="38b2eb", end_color="38b2eb", fill_type="solid")

        sheet_1["A1"].fill = PatternFill(start_color="574fe3", end_color="574fe3", fill_type="solid")
        sheet_1["A1"].font = Font(size=14, color=colors.WHITE, bold=True)

        # inserindo dados na coluna ações
        cont_acoes = 3
        for nome_acao in df_acao["Ação"]:
            sheet_1[f"A{cont_acoes}"] = f"{nome_acao}"
            sheet_1[f"A{cont_acoes}"].fill = PatternFill(start_color="b6e2fa", end_color="b6e2fa", fill_type="solid")
            cont_acoes += 1
        # inserindo dados na coluna quantidade
        cont_quant = 3
        # preparando dados para entrar na planilha
        cot_acao_lista = []
        for k, v in cot_acao.items():
            cot_acao_lista.append(v)

        total_acao_lista = []
        for k, v in total_acao.items():
            total_acao_lista.append(v)

        for quant_acao in df_acao["Quantidade"]:
            cot_acao_real = cot_acao_lista[cont_quant - 3]
            tot_acao_real = total_acao_lista[cont_quant - 3]
            sheet_1[f"B{cont_quant}"] = f"{quant_acao}"
            sheet_1[f"C{cont_quant}"] = f"{cot_acao_real:.2f}"
            sheet_1[f"D{cont_quant}"] = f"{tot_acao_real:.2f}"
            # Colorindo as células da tabela de ação
            sheet_1[f"B{cont_quant}"].fill = PatternFill(start_color="b6e2fa", end_color="b6e2fa", fill_type="solid")
            sheet_1[f"C{cont_quant}"].fill = PatternFill(start_color="b6e2fa", end_color="b6e2fa", fill_type="solid")
            sheet_1[f"D{cont_quant}"].fill = PatternFill(start_color="b6e2fa", end_color="b6e2fa", fill_type="solid")
            cont_quant += 1

        # tabela de moedas
        sheet_1.merge_cells("E1:H1")
        cell_2 = sheet_1.cell(row=1, column=5)
        cell_2.value = "Moeda"
        # Estilizando a tabela de ações
        cell_2.alignment = Alignment(horizontal="center", vertical="center")
        sheet_1["E2"] = "Moeda"
        sheet_1["F2"] = "Quantidade por tipo"
        sheet_1["G2"] = "Cotação (R$)"
        sheet_1["H2"] = "Valor (R$)"

        sheet_1["E2"].fill = PatternFill(start_color="38b2eb", end_color="38b2eb", fill_type="solid")
        sheet_1["F2"].fill = PatternFill(start_color="38b2eb", end_color="38b2eb", fill_type="solid")
        sheet_1["G2"].fill = PatternFill(start_color="38b2eb", end_color="38b2eb", fill_type="solid")
        sheet_1["H2"].fill = PatternFill(start_color="38b2eb", end_color="38b2eb", fill_type="solid")

        sheet_1["E1"].fill = PatternFill(start_color="574fe3", end_color="574fe3", fill_type="solid")
        sheet_1["E1"].font = Font(size=14, color=colors.WHITE, bold=True)

        # inserindo dados na coluna ações
        cont_moeda = 3
        cot_moeda_lista = []
        for k, v in cot_moeda.items():
            cot_moeda_lista.append(v)

        total_moeda_lista = []
        for k, v in total_moeda.items():
            total_moeda_lista.append(v)
        for nome_moeda in df_moeda["Moeda"]:
            sheet_1[f"E{cont_moeda}"] = f"{Funcoes.tire_brl_x(nome_moeda)}"
            sheet_1[f"E{cont_moeda}"].fill = PatternFill(start_color="b6e2fa", end_color="b6e2fa", fill_type="solid")
            cont_moeda += 1
        # inserindo dados na coluna quantidade
        cont_quant_tipo = 3
        for quant_moeda in df_moeda["Quantidade por tipo"]:
            #Arrumando os valores para notação brasileira
            cot_moeda_real = cot_moeda_lista[cont_quant_tipo - 3]
            tot_moeda_real = total_moeda_lista[cont_quant_tipo - 3]
            sheet_1[f"F{cont_quant_tipo}"] = f"{quant_moeda}"
            sheet_1[f"G{cont_quant_tipo}"] = f"{cot_moeda_real:.2f}"
            sheet_1[f"H{cont_quant_tipo}"] = f"{tot_moeda_real:.2f}"
            # Colorindo as células da tabela de ação
            sheet_1[f"F{cont_quant_tipo}"].fill = PatternFill(start_color="b6e2fa", end_color="b6e2fa",
                                                              fill_type="solid")
            sheet_1[f"G{cont_quant_tipo}"].fill = PatternFill(start_color="b6e2fa", end_color="b6e2fa",
                                                              fill_type="solid")
            sheet_1[f"H{cont_quant_tipo}"].fill = PatternFill(start_color="b6e2fa", end_color="b6e2fa",
                                                              fill_type="solid")
            cont_quant_tipo += 1

        # Definindo o tamanho das células
        sheet_1.column_dimensions["A"].width = 15
        sheet_1.column_dimensions["B"].width = 15
        sheet_1.column_dimensions["C"].width = 15
        sheet_1.column_dimensions["D"].width = 15
        sheet_1.column_dimensions["E"].width = 15
        sheet_1.column_dimensions["F"].width = 20
        sheet_1.column_dimensions["G"].width = 15
        sheet_1.column_dimensions["H"].width = 15

        # Salvando a planilha
        planilha = wb.save("planilha_carteira.xlsx")
        Funcoes.colocar_qrcode("planilha_carteira.xlsx")
        return planilha
