import qrcode
import openpyxl
from openpyxl.drawing.image import Image


def soma_valores(nome_arquivo):
    # Abre a planilha
    planilha = openpyxl.load_workbook(nome_arquivo)
    # Abre a folha Carteira
    carteira = planilha["Carteira"]
    # Pega o tamanho da carteira
    tamanho_tabela = len(list(carteira.rows))
    # Define o valor total da carteira
    soma_total = float()
    for linha in range(3, tamanho_tabela+1):
        soma_total += carteira.cell(row=linha, column=4).value
        soma_total += carteira.cell(row=linha, column=8).value
    # Salva a planilha
    return soma_total


def criar_qrcode(nome_arquivo, nome_qrcode):
    # Puxa a função de cima
    soma_total = soma_valores(nome_arquivo)
    # Insere a mensagem que vai ser lida no qrcode
    carteira = (f"O valor da sua carteira é de R$ {round(soma_total, 2)}")
    # Ajusta o tamanho do qrcode
    estilo_qrcode = qrcode.QRCode(box_size=4, border=2)
    # Adiciona amensagem no qrcode
    estilo_qrcode.add_data(carteira)
    # Salva a imagem do qrcode
    imagem = estilo_qrcode.make_image()
    imagem.save(nome_qrcode)


def colocar_qrcode(nome_arquivo, numcarteira):
    # dá nome ao qrcode e salva ele na pasta Imagens
    nome_qrcode = f'Imagens/qrcode{numcarteira}.png'
    criar_qrcode(nome_arquivo, nome_qrcode)
    # Abre a planilha
    planilha = openpyxl.load_workbook(nome_arquivo)
    # Abre a folha Carteira
    carteira = planilha["Carteira"]
    # Pega o tamanho da carteira
    tamanho_tabela = len(list(carteira.rows))
    # Coloca o qrcode aonde é definido
    linha_qrcode = str(tamanho_tabela + 3)
    celula = "D" + linha_qrcode
    imagem = Image(nome_qrcode)
    carteira.add_image(imagem, celula)
    # Salva a planilha
    planilha.save(nome_arquivo)

colocar_qrcode("teste.xlsx", 3)