import openpyxl
from bs4 import BeautifulSoup #Biblioteca para achar tags html, depois vamos colocar num dataframe do pandas
import pandas as pd #Biblioteca para guardar data frame
import requests #Biblioteca para extrair uma URL

def achar_carteira_dic(url_carteira):
    # WebScraping
    # Url do html
    url = url_carteira
    # Obtendo o conteúdo da página em formato de texto

    data = requests.get(url).text
    soup = BeautifulSoup(data, "html.parser")

    # Procurando divs com classes:
    div_acao = soup.find("div", class_="acao")

    # Procurando as tabelas por classe na página
    table_class_acoes = div_acao.find("table")  # tabela de ações

    # Definindo dataframe
    # aqui por enquanto só definimos as colunas
    df_acoes = pd.DataFrame(columns=["Ação", "Quantidade"])  # dataframe da tabela ações

    # Contando o número de linhas em cada tabela
    linhas_na_tabela_acoes = []  # Aqui vamos por um item para linha da tabela com dados, para conseguir o número de linhas
    for row in table_class_acoes.find_all("tr"):  # em html uma linha da tabela é representada pela tag <tr>
        columns_acoes = row.find_all("td")  # em html uma coluna da tabela é representada pela tag <td>
        if (columns_acoes != []):
            linhas_na_tabela_acoes.append("linha")
    num_linhas_acoes = len(linhas_na_tabela_acoes)  # número de linhas em tabela ações
    # Obtendo todas as linhas da tabela
    # Obtendo todas as linhas da tabela acoes

    for row in table_class_acoes.find_all("tr"):  # em html uma linha da tabela é representada pela tag <tr>
        # Obtendo todas as colunas em cada linha
        columns_acoes = row.find_all("td")  # em html uma coluna da tabela é representada pela tag <td>

        if (columns_acoes != []):
            acao = columns_acoes[0].text.strip(" ")
            quantidade = columns_acoes[1].text.strip(" ")
            df_acoes = pd.concat([df_acoes, pd.DataFrame.from_records([{"Ação": acao, "Quantidade": quantidade}])],
                                 ignore_index=True)
            # Aqui a primeira linha vai conter o nosso primeiro df_acoes definido que contém Ação, Quantidade, ou seja os nomes das nossas colunas
            # a segunda parte dos nossos dados serão achados pelo método pandas.DataFrame.from_records
            # nele teremos uma lista é claro, pois em cada linha temos mais de uma informação
            # nesta lista queremos a ação e a quantidade o que a gente tem!
            # pois definimos acao e quantidade
            # Podemos usar o método com dicionário! a chave do dic é a coluna e o valor é o que a coluna vai receber, no caso acao e quantidade IZI
            # o ignore_index é para enumerar nossas linhas, note que ele não coloca numero em ação e quantidade, pois definimos esse dado como coluna

    df_acoes.head(num_linhas_acoes)  # Para retornar as num_linhas_acoes desejadas

    # Refatorando o dataframe
    # Aqui só vou arrumar os dados para ficar da forma certa, pois tratamos eles como str, mas não necessariamente são str
    # Ações
    # Ação
    df_acoes["Ação"] = df_acoes["Ação"].str.upper()
    # Quantidade
    df_acoes["Quantidade"] = [x.replace(',', '.') for x in df_acoes["Quantidade"]]
    df_acoes = df_acoes.astype({"Quantidade": float})

    df_acoes.head(num_linhas_acoes)

    # Procurando divs com classes:
    div_moeda = soup.find("div", class_="moeda")

    # Procurando as tabelas por classe na página
    table_class_moeda = div_moeda.find("table")  # tabela de moeda

    # Definindo dataframe
    # aqui por enquanto só definimos as colunas
    df_moedas = pd.DataFrame(columns=["Moeda", "Quantidade por tipo"])  # dataframe da tabela moeda

    # Contando o número de linhas em cada tabela
    linhas_na_tabela_moeda = []  # Aqui vamos por um item para linha da tabela com dados, para conseguir o número de linhas
    for row in table_class_moeda.find_all("tr"):  # em html uma linha da tabela é representada pela tag <tr>
        columns_moeda = row.find_all("td")  # em html uma coluna da tabela é representada pela tag <td>
        if (columns_moeda != []):
            linhas_na_tabela_moeda.append("linha")
    num_linhas_moeda = len(linhas_na_tabela_moeda)  # número de linhas em tabela ações
    # Obtendo todas as linhas da tabela
    # Obtendo todas as linhas da tabela acoes

    for row in table_class_moeda.find_all("tr"):  # em html uma linha da tabela é representada pela tag <tr>
        # Obtendo todas as colunas em cada linha
        columns_moeda = row.find_all("td")  # em html uma coluna da tabela é representada pela tag <td>

        if (columns_moeda != []):
            moeda = columns_moeda[0].text.strip(" ")
            quantidade_por_tipo = columns_moeda[1].text.strip(" ")
            df_moedas = pd.concat(
                [df_moedas, pd.DataFrame.from_records([{"Moeda": moeda, "Quantidade por tipo": quantidade_por_tipo}])],
                ignore_index=True)
            # Aqui a primeira linha vai conter o nosso primeiro df_acoes definido que contém Moeda, Quantidade por tipo, ou seja os nomes das nossas colunas
            # a segunda parte dos nossos dados serão achados pelo método pandas.DataFrame.from_records
            # nele teremos uma lista é claro, pois em cada linha temos mais de uma informação
            # nesta lista queremos a moeda e a quantidade_por_tipo o que a gente tem!
            # pois definimos moeda e quantidade_por_tipo
            # Podemos usar o método com dicionário! a chave do dic é a coluna e o valor é o que a coluna vai receber, no caso moeda e quantidade_por_tipo IZI
            # o ignore_index é para enumerar nossas linhas, note que ele não coloca numero em moeda e quantidade por tipo, pois definimos esse dado como coluna

    df_moedas.head(num_linhas_moeda)  # Para retornar as num_linhas_acoes desejadas

    # Refatorando o dataframe
    # Aqui só vou arrumar os dados para ficar da forma certa, pois tratamos eles como str, mas não necessariamente são str
    # Ações
    # Ação
    df_moedas["Moeda"] = df_moedas["Moeda"].str.upper()
    # Quantidade
    df_moedas["Quantidade por tipo"] = [x.replace(',', '.') for x in df_moedas["Quantidade por tipo"]]
    df_moedas = df_moedas.astype({"Quantidade por tipo": float})

    df_moedas.head(num_linhas_moeda)

    dic_acao = {}
    cont_dic_acao = 0
    for acao in df_acoes["Ação"]:
        quantidade_acao = df_acoes["Quantidade"][cont_dic_acao]
        dic_acao[f"{acao}"] = f"{quantidade_acao}"
        cont_dic_acao += 1


    dic_moeda = {}
    cont_dic_moeda = 0
    for moeda in df_moedas["Moeda"]:
        quantidade_moeda = df_moedas["Quantidade por tipo"][cont_dic_moeda]
        dic_moeda[f"{moeda}"] = f"{quantidade_moeda}"
        cont_dic_moeda += 1
    dic_moeda_acao = {"Ação": dic_acao, "Moeda": dic_moeda}

    return [dic_moeda_acao, df_moedas, df_acoes]


def gerar_planilha_carteira(url_carteira):
    df_acao = achar_carteira_dic(url_carteira)[2]
    df_moeda = achar_carteira_dic(url_carteira)[1]
    # Criando planilha
    # Criando folha da carteira
    wb = openpyxl.Workbook()
    sheet_1 = wb.active
    sheet_1.title = "Carteira"
    # tabela de ações
    sheet_1["A1"] = "Ações"
    sheet_1.merge_cells("A1:D1")
    sheet_1["A2"] = "Ação"
    sheet_1["B2"] = "Quantidade"
    sheet_1["C2"] = "Cotação da ação"
    sheet_1["D2"] = "Valor atual total ação"
    # inserindo dados na coluna ações
    cont_acoes = 3
    for nome_acao in df_acao["Ação"]:
        sheet_1[f"A{cont_acoes}"] = f"{nome_acao}"
        #pintar essa célula sheet_1[f"A{cont_acoes}"]
        cont_acoes += 1
    # inserindo dados na coluna quantidade
    cont_quant = 3
    for quant_acao in df_acao["Quantidade"]:
        sheet_1[f"B{cont_quant}"] = f"{quant_acao}"
        sheet_1[f"C{cont_quant}"] = f"{1}"
        sheet_1[f"D{cont_quant}"] = f"{3}"
        cont_quant += 1

    # tabela de moedas
    sheet_1["E1"] = "Moedas"
    sheet_1.merge_cells("E1:H1")
    sheet_1["E2"] = "Moeda"
    sheet_1["F2"] = "Quantidade por tipo"
    sheet_1["G2"] = "Cotação"
    sheet_1["H2"] = "Valor"
    # inserindo dados na coluna ações
    cont_moeda = 3
    for nome_moeda in df_moeda["Moeda"]:
        sheet_1[f"E{cont_moeda}"] = f"{nome_moeda}"
        cont_moeda += 1
    # inserindo dados na coluna quantidade
    cont_quant_tipo = 3
    for quant_moeda in df_moeda["Quantidade por tipo"]:
        sheet_1[f"F{cont_quant_tipo}"] = f"{quant_moeda}"
        sheet_1[f"G{cont_quant_tipo}"] = f"{1}"
        sheet_1[f"H{cont_quant_tipo}"] = f"{3}"
        cont_quant_tipo += 1
    # Salvando planilha
    planilha = wb.save("planilha_carteira.xlsx")
    return planilha

gerar_planilha_carteira("https://anajuliarlc.github.io/Trabalho-de-IC/sitecar/carteira%201")
